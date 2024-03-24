import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from Gui2_5 import *


class Excelopener:
    def __init__(self, filename, max_row, min_row):
        self.max_row = max_row + 1
        self.min_row = min_row + 1
        self.pd_table = None
        self.pd_table2 = None
        if filename:
            self.book = openpyxl.open(filename, read_only=True)
            self.sheet = self.book.active
            self.data = []
            for row in self.sheet.iter_rows(values_only=True):
                self.data.append(row)

    def create_a_table_in_py(self):
        self.pd_table = pd.DataFrame(self.data[self.min_row:self.max_row], columns=self.data[0])
        return self.pd_table

    def filter_df(self, criterion, criterion2):
        a = []
        self.pd_table2 = self.pd_table.set_index(criterion)
        for i in self.pd_table2.index:
            if i != criterion2:
                a.append(i)
        return self.pd_table2.drop(a, axis=0)


class Create_plot:
    def __init__(self, number_of_colum_1, number_of_colum_2, max_row,
                 min_row, tabel, type_of_plot="plot", color="blue"):
        self.max_row = max_row
        self.min_row = min_row
        self.table = Excelopener(tabel, max_row, min_row).create_a_table_in_py()
        self.column_1 = self.table[number_of_colum_1]
        self.column_2 = self.table[number_of_colum_2]
        a = [column for column in self.table]
        self.x_label = a[0]
        self.y_label = a[1]
        self.ax = plt.gca()

        self.plot_type = type_of_plot
        self.color = color
        self.count = 2
        self.all_plots = {1: [self.column_1, self.column_2, type_of_plot, self.color]}

        print("create: completed")

    def plot_data(self):
        if self.plot_type == "plot":
            self.ax.plot(self.column_1, self.column_2, color=self.color)
        elif self.plot_type == "bar":
            self.ax.bar(self.column_1, self.column_2, color=self.color)
        elif self.plot_type == "scatter":
            self.ax.scatter(self.column_1, self.column_2, color=self.color)
        elif self.plot_type == "stem":
            self.ax.stem(self.column_1, self.column_2, linefmt=self.color, markerfmt=self.color, basefmt='black')
        elif self.plot_type == "step":
            self.ax.step(self.column_1, self.column_2, where='mid', color=self.color)
        print("ploting: completed")

    def set_labels(self, x_label, y_label):
        self.ax.set_xlabel(x_label)
        self.ax.set_ylabel(y_label)
        print("labels are set")

    def set_title(self, title):
        self.ax.set_title(title)
        print("title is set")

    def add_new_plot(self, column_3, column_4, plot_type="plot", color="orange"):
        self.all_plots[self.count] = [column_3, column_4, plot_type, color]
        self.count += 1
        if plot_type == "plot":
            self.ax.plot(self.table[column_3], self.table[column_4], color=color)
        elif plot_type == "bar":
            self.ax.bar(self.table[column_3], self.table[column_4], color=color)
        elif plot_type == "scatter":
            self.ax.scatter(self.table[column_3], self.table[column_4], color=color)
        elif plot_type == "stem":
            self.ax.stem(self.table[column_3], self.table[column_4], linefmt=color, markerfmt=color, basefmt='black')
        elif plot_type == "step":
            self.ax.step(self.table[column_3], self.table[column_4], where='mid', color=color)

    def show_plot(self):
        plt.show()
        print("show: completed")

    def change_plot_type(self, index, plot_type):
        self.ax.clear()
        for i, j in self.all_plots.items():
            if i == index:
                if plot_type == "plot":
                    self.ax.plot(j[0], j[1], color=j[-1])
                elif plot_type == "bar":
                    self.ax.bar(j[0], j[1], color=j[-1])
                elif plot_type == "scatter":
                    self.ax.scatter(j[0], j[1], color=j[-1])
                elif plot_type == "stem":
                    self.ax.stem(j[0], j[1], linefmt=j[-1], markerfmt=j[-1], basefmt='black')
                elif plot_type == "step":
                    self.ax.step(j[0], j[1], where='mid', color=j[-1])
            else:
                if j[2] == "plot":
                    self.ax.plot(j[0], j[1], color=j[-1])
                elif j[2] == "bar":
                    self.ax.bar(j[0], j[1], color=j[-1])
                elif j[2] == "scatter":
                    self.ax.scatter(j[0], j[1], color=j[-1])
                elif j[2] == "stem":
                    self.ax.stem(j[0], j[1], linefmt=j[-1], markerfmt=j[-1], basefmt='black')
                elif j[2] == "step":
                    self.ax.step(j[0], j[1], where='mid', color=j[-1])
        print("plot is changed")

    def delete_plot(self, index):
        self.ax.clear()
        if index in self.all_plots:
            del self.all_plots[index]
        for i, j in self.all_plots.items():
            if j[2] == "plot":
                self.ax.plot(j[0], j[1], color=j[-1])
            elif j[2] == "bar":
                self.ax.bar(j[0], j[1], color=j[-1])
            elif j[2] == "scatter":
                self.ax.scatter(j[0], j[1], color=j[-1])
            elif j[2] == "stem":
                self.ax.stem(j[0], j[1], linefmt=j[-1], markerfmt=j[-1], basefmt='black')
            elif j[2] == "step":
                self.ax.step(j[0], j[1], where='mid', color=j[-1])
        print("plot is deleted")

    def change_color(self, index, color):
        self.ax.clear()
        for i, j in self.all_plots.items():
            if i == index:
                if j[2] == "plot":
                    self.ax.plot(j[0], j[1], color=color)
                elif j[2] == "bar":
                    self.ax.bar(j[0], j[1], color=color)
                elif j[2] == "scatter":
                    self.ax.scatter(j[0], j[1], color=color)
                elif j[2] == "stem":
                    self.ax.stem(j[0], j[1], linefmt=color, markerfmt=color, basefmt='black')
                elif j[2] == "step":
                    self.ax.step(j[0], j[1], where='mid', color=color)
            else:
                if j[2] == "plot":
                    self.ax.plot(j[0], j[1], color=j[-1])
                elif j[2] == "bar":
                    self.ax.bar(j[0], j[1], color=j[-1])
                elif j[2] == "scatter":
                    self.ax.scatter(j[0], j[1], color=j[-1])
                elif j[2] == "stem":
                    self.ax.stem(j[0], j[1], linefmt=j[-1], markerfmt=j[-1], basefmt='black')
                elif j[2] == "step":
                    self.ax.step(j[0], j[1], where='mid', color=j[-1])

    def save_plot(self, name_of_plot):
        self.ax.savefig(name_of_plot)
        print("plot is save")

    def create_pie_plot(self, column_5, label):
        self.ax.pie(self.table[column_5], labels=self.table[label])

    def set_legend(self, data):
        self.ax.legend(data)


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())
